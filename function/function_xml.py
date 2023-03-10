import xml.etree.cElementTree as ET
import openpyxl
from openpyxl.styles import PatternFill
from database.raw_data_string import *
import function.function_compare as fc_cp


redFill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
## read xslx file database CE

class xml_para():
    
    def __init__(self, guid, name, key_10 ):
        self.__guid = guid
        self.__name = name
        self.__key_10 = key_10
    
    def get_val(self):
        return self.__key_10

class tool_value():
    def __init__(self, item, value):
        self.__item = item
        self.__value = value
        
    def gel_val(self):
        return self.value

## Parse xml file to object
def parse_para_xml(file_name, worksheet):
    parsed = ET.parse(file_name)
    root = parsed.getroot()
    List = []
    
    for step in root:
        for i in range(7, size):
            if step.attrib.get('guid') == worksheet.cell(row=i, column=20).value:
                worksheet.cell(row=i, column=7).value = step[1].text
                List.append(xml_para(step.attrib.get("guid"), step.attrib.get("name"), step[1].text))
            elif step.attrib.get('name') == worksheet.cell(row=i, column=15).value:
                worksheet.cell(row=i, column=7).value = step[1].text
                List.append(xml_para(step.attrib.get("guid"), step.attrib.get("name"), step[1].text))

def parse_para_xml_db(file_name, df):
    parsed = ET.parse(file_name)
    root = parsed.getroot()
    List = []
    tool_values = []
    for step in root:
        for i in range(size_df):
            if step.attrib.get('guid') == df.iloc[i,6]:
                df.iloc[i,7] = step[1].text
                List.append(xml_para(step.attrib.get("guid"), step.attrib.get("name"), step[1].text))
            elif step.attrib.get('name') == df.iloc[i,5]:
                df.iloc[i,7] = step[1].text
                List.append(xml_para(step.attrib.get("guid"), step.attrib.get("name"), step[1].text))
            
def color_cell(worksheet):
    for i in range(7, size):
        if fc_cp.compare(worksheet.cell(row =i, column =7), worksheet.cell(row =i, column =5), worksheet.cell(row =i, column =6)) == 0:
            worksheet.cell(row=i, column=7).fill = redFill